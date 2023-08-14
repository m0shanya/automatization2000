from dotenv import load_dotenv, find_dotenv
from copy import deepcopy
from openpyxl import styles, load_workbook

import os
import fdb
import pandas as pd
import datetime as dt

load_dotenv(find_dotenv())
FB_HOST = os.environ.get("FB_HOST")
FB_DATABASE = os.environ.get("FB_DATABASE")
FB_USER = os.environ.get("FB_USER")
FB_PASSWORD = os.environ.get("FB_PASSWORD")
CHARSET = os.environ.get("CHARSET")

connection = {'host': FB_HOST,
              'database': FB_DATABASE,
              'user': FB_USER,
              'password': FB_PASSWORD,
              'charset': CHARSET
              }

time_dct = {"time": ["00:00-00:30", "00:30-01:00", "01:00-01:30", "01:30-02:00", "02:00-02:30", "02:30-03:00",
                     "03:00-03:30", "03:30-04:00", "04:00-04:30", "04:30-05:00", "05:00-05:30", "05:30-06:00",
                     "06:00-06:30", "06:30-07:00", "07:00-07:30", "07:30-08:00", "08:00-08:30", "08:30-09:00",
                     "09:00-09:30", "09:30-10:00", "10:00-10:30", "10:30-11:00", "11:00-11:30", "11:30-12:00",
                     "12:00-12:30", "12:30-13:00", "13:00-13:30", "13:30-14:00", "14:00-14:30", "14:30-15:00",
                     "15:00-15:30", "15:30-16:00", "16:00-16:30", "16:30-17:00", "17:00-17:30", "17:30-18:00",
                     "18:00-18:30", "18:30-19:00", "19:00-19:30", "19:30-20:00", "20:00-20:30", "20:30-21:00",
                     "21:00-21:30", "21:30-22:00", "22:00-22:30", "22:30-23:00", "23:00-23:30", "23:30-24:00"]}

ku = 100


def changer_command(command_name: str) -> int:
    cmdid = 0
    if command_name == "Срез 30 мин E+":
        cmdid = 13
    elif command_name == "Срез 30 мин E-":
        cmdid = 14
    elif command_name == "Срез 30 мин R+":
        cmdid = 15
    else:
        cmdid = 16
    return cmdid


def get_data(numb: list, time: list, cmd_name: str) -> dict:
    dlc = ["-"]
    res = {}

    command_id = changer_command(cmd_name)

    with fdb.connect(**connection) as con:
        for element in numb:
            dates = get_date(time)

            value = []

            cur = con.cursor()

            query_element = f"""SELECT M_SWVMID FROM SL3VMETERTAG 
            WHERE M_SVMETERNAME='{element}'"""
            cur.execute(query_element)
            vmid = cur.fetchone()

            query = f"""SELECT * FROM L2HALF_HOURLY_ENERGY 
            WHERE M_SWCMDID={command_id} AND 
            M_SDTDATE BETWEEN \'{time[0]}\' and \'{time[1]}\' AND M_SWVMID={vmid[0]} ORDER BY M_SDTDATE"""
            cur.execute(query)
            i = 0
            j = 0
            for val in cur.fetchall():

                for item in dates:
                    if i - j > 0:
                        item = dates[i - j]
                    i += 1
                    tmp = item.split("-")
                    if dt.datetime(int(tmp[0]), int(tmp[1]), int(tmp[2])) == val[5]:
                        j += 1
                        dates.remove(item)
                        value += [number * ku if number is not None else number for number in val[6:54]]
                        break
                    else:
                        value = value + dlc * 48
                        item = dates[i - j]

            res[f"Датчик {element}"] = value

    return res


def get_date(time: list) -> list:
    start_date = time[0]
    end_date = time[1]
    dates = pd.date_range(
        min(start_date, end_date),
        max(start_date, end_date)
    ).strftime('%Y-%m-%d').tolist()

    return dates


def do_write(val: dict, list_of_dates: list, vmids: list, filename: str, t_start: str, t_end: str):
    prime_values = dict(sorted(security(val, list_of_dates, t_start, t_end).items()))
    dataframe = pd.DataFrame(prime_values)

    with pd.ExcelWriter(f'{filename}.xlsx', engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=f'{vmids[0]}..{vmids[-1]}',
                           index=False)

    max_min_func(dataframe, filename, vmids)


def max_min_func(df: pd.DataFrame, name: str, vm_ids: list):
    maximum = 0
    minimum = 0

    for val in df.columns:
        if "Датчик" in val:
            maximum = df[val].max()
            minimum = df[val].min()
            max_idx = df.loc[df[val] == maximum].index[0]
            min_idx = df.loc[df[val] == minimum].index[0]
            col_idx = df.columns.get_loc(val)

            painter(name, vm_ids, max_idx, min_idx, col_idx)


def painter(file_name: str, sheet_name: list, *args):
    wb = load_workbook(f"{file_name}.xlsx")
    ws = wb.get_sheet_by_name(f"{sheet_name[0]}..{sheet_name[-1]}")
    ws.cell(row=args[0] + 2, column=args[2] + 1).fill = styles.PatternFill(start_color='3aa832', end_color='3aa832',
                                                                           fill_type='solid')
    ws.cell(row=args[1] + 2, column=args[2] + 1).fill = styles.PatternFill(start_color='51fa0e62', end_color='51fa0e62',
                                                                           fill_type='solid')
    wb.save(f"{file_name}.xlsx")


def time_manager(time_s: str, time_e: str) -> list[int]:
    index_time = []
    if (time_s in time_dct["time"]) and (time_e in time_dct["time"]):
        index_time.append(time_dct["time"].index(time_s))
        index_time.append(time_dct["time"].index(time_e))
    return index_time


def time_dct_editor(idx: list[int], len_of_frame: int):
    minus = ["-"]
    tmp_res = []
    tmp_time = deepcopy(time_dct["time"])
    tmp_time[0:idx[0]] = minus * idx[0]
    tmp_res = tmp_res + tmp_time
    tmp_res = tmp_res + time_dct["time"] * int((len_of_frame / 48) - 2)
    tmp_time = deepcopy(time_dct["time"])
    tmp_time[idx[1] + 1:len(tmp_time)] = minus * (47 - idx[1])
    tmp_res = tmp_res + tmp_time

    return tmp_res


def security(dct_of_values: dict, spisok_dat: list, start_time: str, end_time: str):
    spis_of_dates = []
    for item in spisok_dat:
        spis_of_dates = spis_of_dates + [item] * 48

    dct_of_values["date"] = spis_of_dates
    max_len_list = 0
    for val in dct_of_values.values():
        max_len_list = len(val)
        break

    for val in dct_of_values.values():
        max_len = len(val)
        if max_len > max_len_list:
            max_len_list = max_len

    dlc = ["-"]

    for key in dct_of_values.keys():
        if len(dct_of_values[key]) < max_len_list:
            dct_of_values[key] = dct_of_values[key] + dlc * (max_len_list - len(dct_of_values[key]))

    # dct_of_values["time"] = time_dct["time"] * int((max_len_list / 48))
    dct_of_values["time"] = time_dct_editor(time_manager(start_time, end_time), max_len_list)

    return dct_of_values


if __name__ == "__main__":
    time_list = ['2023-04-26 00:00:00', '2023-04-28 00:00:00']
    vmid_list = ['СТП-1 КВ-1', 'ПП-3 ТО-1']
    dates = get_date(time_list)
    values = get_data(vmid_list, time_list, 'Срез 30 мин E+')
    do_write(values, dates, vmid_list, 'Срез 30 мин E+', "01:00-01:30", "06:00-06:30")
