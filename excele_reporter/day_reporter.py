from copy import deepcopy
from dotenv import load_dotenv, find_dotenv
from openpyxl import styles, load_workbook

import os
import fdb
import pandas as pd

load_dotenv(find_dotenv())
FB_HOST = os.environ.get("FB_HOST")
FB_DATABASE = os.environ.get("FB_DATABASE")
FB_USER = os.environ.get("FB_USER")
FB_PASSWORD = os.environ.get("FB_PASSWORD")
CHARSET = os.environ.get("CHARSET")

connection = {'host': FB_HOST,
              'database': FB_DATABASE,
              'user': FB_USER,
              'password': FB_PASSWORD,
              'charset': CHARSET
              }

tarif = {"tarif": [0, 1, 2, 3, 4, 5, "Сумма"]}

ku = 4000


def changer_command_day(command_name: str) -> int:
    cmdid = 0
    if command_name == "Начало суток E+":
        cmdid = 17
    elif command_name == "Начало суток E-":
        cmdid = 18
    elif command_name == "Начало суток R+":
        cmdid = 19
    else:
        cmdid = 20
    return cmdid


def get_day_data(numb: list, time: list, cmd_name: str) -> dict:
    dlc = ["-"]
    res = {}

    command_id = changer_command_day(cmd_name)

    with fdb.connect(**connection) as con:
        for element in numb:
            dates = get_day_date(time)

            value = []

            cur = con.cursor()

            query_element = f"""SELECT M_SWVMID FROM SL3VMETERTAG 
            WHERE M_SVMETERNAME='{element}'"""
            cur.execute(query_element)
            vmid = cur.fetchone()

            query = f"""SELECT DISTINCT M_SWTID, M_STIME, M_SFVALUE FROM L3ARCHDATA 
            WHERE M_SWCMDID={command_id} AND 
            M_STIME BETWEEN \'{time[0]}\' and \'{time[1]}\' AND M_SWVMID={vmid[0]} ORDER BY M_STIME"""
            cur.execute(query)

            tmp = False
            tmp_list = []
            tmp_sum = 0
            trf = 0
            res[f"Счётчик {element}"] = []
            for val in cur.fetchall():
                for item in dates:

                    if tmp:
                        item = tmp
                        if tmp < val[1].strftime('%Y-%m-%d') and len(res[f"Счётчик {element}"]) % 7 != 0:
                            for _ in range(len(res[f"Счётчик {element}"]),
                                           len(tarif["tarif"]) * (dates.index(tmp) + 1) - 1):
                                res[f"Счётчик {element}"].append('-')
                            res[f"Счётчик {element}"].append(tmp_sum)
                            tmp_sum = 0
                            tmp = False

                    if item == val[1].strftime('%Y-%m-%d') and trf == val[0]:
                        if item not in tmp_list:
                            tmp_list.append(item)
                        trf += 1
                        tmp = item
                        res[f"Счётчик {element}"] += [val[2] if val[2] is not None else val[2]]
                        tmp_sum += val[2] / ku if val[2] is not None else 0
                        break
                    elif item == val[1].strftime('%Y-%m-%d') and trf != val[0]:
                        trf = 0
                        tmp = item
                        trf += 1
                        res[f"Счётчик {element}"] += [val[2] if val[2] is not None else val[2]]
                        tmp_sum += val[2] / ku if val[2] is not None else 0
                        break
                    elif item not in tmp_list:
                        tmp_list.append(item)
                        trf = 0
                        res[f"Счётчик {element}"] = res[f"Счётчик {element}"] + dlc * 7

            if tmp:
                for _ in range(len(res[f"Счётчик {element}"]),
                               len(tarif["tarif"]) * (dates.index(tmp) + 1) - 1):
                    res[f"Счётчик {element}"].append('-')
                res[f"Счётчик {element}"].append(tmp_sum)
                tmp_sum = 0


    return res


def get_day_date(time: list) -> list:
    start_date = time[0]
    end_date = time[1]
    dates = pd.date_range(
        min(start_date, end_date),
        max(start_date, end_date)
    ).strftime('%Y-%m-%d').tolist()

    return dates


def do_day_write(val: dict, list_of_dates: list, vmids: list, filename: str):
    prime_values = dict(sorted(security(val, list_of_dates).items()))
    dataframe = pd.DataFrame(prime_values)

    with pd.ExcelWriter(f'days/{filename}.xlsx', engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=f'{vmids[0]}..{vmids[-1]}',
                           index=False)
    max_min_func_for_day(dataframe, filename, vmids)


def max_min_func_for_day(df: pd.DataFrame, name: str, vm_ids: list):
    maximum = 0
    minimum = 0

    for val in df.columns:
        if "Счётчик" in val:
            number = []

            for element in df[val]:
                if element != '-':
                    number.append(element)
            if not number:
                continue

            maximum = max(number)
            minimum = min(number)

            max_idx = df.loc[df[val] == maximum].index[0]
            min_idx = df.loc[df[val] == minimum].index[0]
            col_idx = df.columns.get_loc(val)

            painter_for_day(name, vm_ids, max_idx, min_idx, col_idx)


def painter_for_day(file_name: str, sheet_name: list, *args):
    wb = load_workbook(f"days/{file_name}.xlsx")
    ws = wb.get_sheet_by_name(f"{sheet_name[0]}..{sheet_name[-1]}")
    ws.cell(row=args[0] + 2, column=args[2] + 1).fill = styles.PatternFill(start_color='3aa832', end_color='3aa832',
                                                                           fill_type='solid')
    ws.cell(row=args[1] + 2, column=args[2] + 1).fill = styles.PatternFill(start_color='51fa0e62', end_color='51fa0e62',
                                                                           fill_type='solid')
    wb.save(f"days/{file_name}.xlsx")


def security(dct_of_values: dict, spisok_dat: list):
    spis_of_dates = []
    for item in spisok_dat:
        spis_of_dates = spis_of_dates + [item] * 7

    dct_of_values["date"] = spis_of_dates
    max_len_list = 0
    for val in dct_of_values.values():
        max_len_list = len(val)
        break

    for val in dct_of_values.values():
        max_len = len(val)
        if max_len > max_len_list:
            max_len_list = max_len

    dlc = ["-"]

    for key in dct_of_values.keys():
        if len(dct_of_values[key]) < max_len_list:
            dct_of_values[key] = dct_of_values[key] + dlc * (max_len_list - len(dct_of_values[key]))

    dct_of_values["Tarif"] = tarif["tarif"] * int(max_len_list / len(tarif["tarif"]))

    return dct_of_values


if __name__ == "__main__":
    time_list = ['2022-12-17 00:00:00', '2023-02-28 00:00:00']
    vmid_list = ['Office SS301']
    dates = get_day_date(time_list)
    values = get_day_data(vmid_list, time_list, 'Начало суток E+')
    do_day_write(values, dates, vmid_list, 'Начало суток E+')
